"""
time: 2020/7/3
author: best_yzw
aim: 逐条检测网盘分享链接的可用性，如果不可用，输出相关提示
"""

# 网黑哥资源合集： https://mp.weixin.qq.com/s/hPgmL2WOU9iJWAqnl8CMsQ

# 某资源
# 百度云： https://pan.baidu.com/s/1PDxlvzYu9QOaMOY7ekFi1Q
# 密码： d933
# 蓝奏云： https://www.lanzous.com/b384810/
# 密码： 41rg

# 百度云失效链接： https://pan.baidu.com/share/link?uk=1813251526&shareid=540167442
# https://pan.baidu.com/s/13Ljkj-pYJmoh0VKLGvg_GQ

from selenium import webdriver
import time
import re
import openpyxl
import os


def parser_url(link):
    """
    :param link: 百度网盘下载链接: https://pan.baidu.com/s/13Ljkj-pYJmoh0VKLGvg_GQ 提取码: 2333
    :return: https://pan.baidu.com/s/13Ljkj-pYJmoh0VKLGvg_GQ
    """
    try:
        return re.search(r'https?://pan\.baidu\.com/s/.{23}', link.strip()).group()
    except:
        print(' 错误： 未找到链接')
        return None


def check_url(url):
    ops = webdriver.ChromeOptions()
    ops.add_argument('--headless')
    ops.add_argument('--disable-gpu')
    browser = webdriver.Chrome(options=ops)
    browser.get(url)
    html = browser.page_source
    browser.close()
    # print(html)
    if html.find('给您加密分享了文件') == -1:
        # print(url, '失效')
        return 0
    else:
        # print(url, '可用')
        return 1


def check_data(data):
    status = []
    for record in data:
        start_time = time.time()
        print('正在检测 ' + record[0] + ' ...', end='')
        url = parser_url(record[1])
        if url == None:
            print('目前只支持检测：百度网盘的分享链接')
            print('如： "https://pan.baidu.com/s/1nsWg3EAGFzxRhaEhb5bxvg" ')
            print(record[0] +' 将被标记为 "未检测", 检测继续...')
            status.append('未检测')
            continue
        if check_url(url):
            status.append('可用')
            print(' 可用  ', end='')
        else:
            status.append('失效')
            print(' 失效  ', end='')
        end_time = time.time()
        print('耗时：'+str(round((end_time - start_time), 2)) + 's')
    return status

if __name__ == '__main__':

    welcome = '''
__________                 __    _____.___.                
\______   \ ____   _______/  |_  \__  |   |_________  _  __
 |    |  _// __ \ /  ___/\   __\  /   |   \___   /\ \/ \/ /
 |    |   \  ___/ \___ \  |  |    \____   |/    /  \     / 
 |______  /\___  >____  > |__|    / ______/_____ \  \/\_/  
        \/     \/     \/          \/            \/                 
    '''

    print(welcome)
    print('欢迎使用 《链接失效否？》，一个批量判断分享链接是否失效的小工具。')
    print('程序最后调试日期为2020/7/4，调试无误。')
    print()
    print('请输入文件名 （如： links.xlsx 或 D:\links.xlsx）')
    filename = input()

    # 读取 links.xlsx
    # 得到 [[资源描述，资源链接，状态]，
    #       [资源描述，资源链接，状态]，
    #       ...
    # ]

    # 判断文件是否存在
    if not os.path.exists(filename):
        print('错误：' + filename +' 不存在')
        input("按 enter 退出")
        os._exit(0)

    # 打开文件，获取工作簿对象
    try:
        work_book = openpyxl.load_workbook(filename)
    except:
        print("错误：无法正常打开 " + filename)
        print("请确保：")
        print("\t1." + filename +" 类型为 xlsx")
        print("\t2." + filename +" 用任意表格软件，如：excel 能正常打开查看")
        input("按 enter 退出")
        os._exit(0)


    # 统计数据
    start_time = time.time()
    sum_tests = 0
    broken_tests = 0
    cannot_tests = 0

    print('由于程序特性，检测一条链接大概需要5秒，所以现在您可以去放松一下，稍后再来。')
    for sheet_name in work_book.sheetnames:
        data = []
        work_sheet = work_book[sheet_name]
        for row in list(work_sheet.rows)[1:]:
            row_data = []
            for cell in row[:3]:
                row_data.append(cell.value)
            data.append(row_data)

        sum_tests += len(data)
        # 逐一检测链接，并修改链接状态
        result = check_data(data)
        for i in range(len(result)):
            if result[i] == '失效':
                work_sheet.cell(i+2, 3, result[i])
                broken_tests += 1
            elif result[i] == '未检测':
                work_sheet.cell(i+2, 3, result[i])
                cannot_tests += 1
            else:
                work_sheet.cell(i+2, 3, '')

    # 保存检测结果
    save_name = filename[:-5] + '(已检测).xlsx'
    work_book.save(save_name)
    work_book.close()

    print()
    print()
    print('检测完毕！')
    print('共检测 ' + str(sum_tests) + ' 条链接，用时 ' + str(round((time.time() - start_time), 2)) + 's')
    print('发现失效链接 ' + str(broken_tests) + ' 条，程序未能正常检测链接 ' + str(cannot_tests) + ' 条')
    print('详细检测结果已保存到 "' + save_name + '"')

